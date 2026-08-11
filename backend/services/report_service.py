from models import db, DentalEntry, Doctor, Hospital
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import os
from datetime import datetime

def generate_excel_report(date_from, date_to, doctor_id=None, hospital_id=None):
    """Generate Excel report with filters and return file path."""
    query = DentalEntry.query.filter(
        DentalEntry.entry_date.between(date_from, date_to)
    )
    if doctor_id:
        query = query.filter_by(doctor_id=doctor_id)
    if hospital_id:
        query = query.filter_by(hospital_id=hospital_id)
    
    entries = query.order_by(DentalEntry.entry_date).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue Report"
    
    # Headers
    headers = ['Entry No', 'Date', 'Doctor', 'Hospital', 'Patient', 'Description',
               'Units', 'Shade', 'Work Type', 'Amount']
    ws.append(headers)
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for e in entries:
        ws.append([
            e.entry_no,
            e.entry_date.strftime('%Y-%m-%d'),
            e.doctor.doctor_name,
            e.hospital.hospital_name,
            e.patient_name,
            e.description,
            e.no_of_units,
            e.shade_type,
            e.work_type,
            float(e.amount)
        ])
    
    # Add totals
    total_row = len(entries) + 2
    ws[f'A{total_row}'] = "GRAND TOTAL"
    ws[f'J{total_row}'] = f"=SUM(J2:J{total_row-1})"
    ws[f'A{total_row}'].font = Font(bold=True)
    
    # Auto column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save file
    os.makedirs('temp_excel', exist_ok=True)
    filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join('temp_excel', filename)
    wb.save(filepath)
    return filepath